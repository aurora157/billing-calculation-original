import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def sanitize_sheet_name(name):
    # Remove invalid characters and truncate to a maximum of 31 characters
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in invalid_chars:
        name = name.replace(ch, '')
    return name[:31]

def write_dataframe_to_sheet(ws, df, header):
    """
    Writes the DataFrame 'df' into worksheet 'ws' starting at row 2,
    using the provided header list (assumed to be in row 1).
    """
    # Clear any existing data below the header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    # Write each row from the DataFrame starting at row 2
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

def apply_header_formatting(ws, num_columns):
    # Define header styling: blue fill, white bold centered text, thin borders.
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_adjust_column_widths(ws):
    # Auto-adjust column widths based on the maximum length of cell values.
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

def process_icr_report(services_file: str, usage_file: str, client_name: str) -> str:
    """
    Process ICR report from usage and services CSV files.
    
    Args:
        services_file (str): Path to the services breakdown CSV file
        usage_file (str): Path to the usage CSV file
        client_name (str): Name of the client for the output file
        
    Returns:
        str: Path to the generated report
    """
    # Create a new workbook
    from openpyxl import Workbook
    wb = Workbook()
    
    # Load CSV files into DataFrames
    try:
        usage_df = pd.read_csv(usage_file)
        print(f"Loaded usage data: {usage_file} (rows: {len(usage_df)})")
    except Exception as e:
        print(f"Error loading usage file: {e}")
        raise Exception(f"Failed to load usage file: {e}")
    
    try:
        services_df = pd.read_csv(services_file)
        print(f"Loaded services breakdown: {services_file} (rows: {len(services_df)})")
    except Exception as e:
        print(f"Error loading services breakdown file: {e}")
        raise Exception(f"Failed to load services breakdown file: {e}")

    # Process the "Usage Charges" column:
    # Remove currency symbols (like "£") and commas, then convert to float.
    if 'usage charges' in services_df.columns:
        services_df['usage charges'] = services_df['usage charges'].replace({r'[£,]': ''}, regex=True).astype(float)
    elif 'Usage Charges' in services_df.columns:
        services_df['Usage Charges'] = services_df['Usage Charges'].replace({r'[£,]': ''}, regex=True).astype(float)
        services_df.rename(columns={"Usage Charges": "usage charges"}, inplace=True)
    else:
        print("Warning: 'Usage Charges' column not found in the services breakdown file.")
    
    # Filter for rows where Usage Charges are above £5.00
    filtered_services = services_df[services_df['usage charges'] > 5.00]
    print(f"Found {len(filtered_services)} rows in services breakdown with Usage Charges above £5.00.")

    # Create the main sheet "ICR Bill Summary"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])  # Remove default sheet
    main_sheet = wb.create_sheet("ICR Bill Summary")
    print("Using sheet 'ICR Bill Summary' for usage data.")

    # Use columns from the usage file as header
    header = list(usage_df.columns)
    for col_num, col_name in enumerate(header, start=1):
        main_sheet.cell(row=1, column=col_num, value=col_name)

    # Write the entire usage data to the "ICR Bill Summary" sheet (starting at row 2)
    write_dataframe_to_sheet(main_sheet, usage_df, header)
    # Apply header formatting and auto-adjust column widths on the main sheet
    apply_header_formatting(main_sheet, len(header))
    auto_adjust_column_widths(main_sheet)
    print("Wrote usage data into 'ICR Bill Summary' sheet with formatting.")

    # Create new sheets for each qualifying service from the services breakdown.
    if 'Service' in filtered_services.columns:
        unique_services = filtered_services['Service'].unique()
        for service in unique_services:
            # Filter usage data for the current service
            service_usage = usage_df[usage_df['Service'] == service]
            if service_usage.empty:
                print(f"No usage records found for Service: {service}")
                continue
            
            # If the usage data has a UserName column, create separate sheets per user
            if 'UserName' in usage_df.columns:
                unique_users = service_usage['UserName'].unique()
                for user in unique_users:
                    sheet_name = sanitize_sheet_name(f"{service} - {user}")
                    new_sheet = wb.create_sheet(sheet_name)
                    # Write header row in the new sheet
                    for col_num, col_name in enumerate(header, start=1):
                        new_sheet.cell(row=1, column=col_num, value=col_name)
                    # Filter usage data for both the service and the user
                    filtered_usage = service_usage[service_usage['UserName'] == user]
                    if filtered_usage.empty:
                        print(f"No usage records for Service: {service} and User: {user}")
                        continue
                    write_dataframe_to_sheet(new_sheet, filtered_usage, header)
                    apply_header_formatting(new_sheet, len(header))
                    auto_adjust_column_widths(new_sheet)
                    print(f"Created sheet '{sheet_name}' with {len(filtered_usage)} rows and formatted header.")
            else:
                # If no UserName column exists, create one sheet per service.
                sheet_name = sanitize_sheet_name(service)
                new_sheet = wb.create_sheet(sheet_name)
                for col_num, col_name in enumerate(header, start=1):
                    new_sheet.cell(row=1, column=col_num, value=col_name)
                write_dataframe_to_sheet(new_sheet, service_usage, header)
                apply_header_formatting(new_sheet, len(header))
                auto_adjust_column_widths(new_sheet)
                print(f"Created sheet '{sheet_name}' with {len(service_usage)} rows and formatted header.")
    else:
        print("Skipping additional sheets; 'Service' column not found in the services breakdown file.")

    # Save to Excel in Downloads folder with month_year_Bill_Run format
    downloads_path = os.path.expanduser("~/Downloads")
    prev_month = (pd.Timestamp.now() - pd.DateOffset(months=1)).strftime('%B')
    current_year = pd.Timestamp.now().year
    output_path = os.path.join(downloads_path, f"{prev_month}_{current_year}_Bill_Run_{client_name}.xlsx")
    try:
        wb.save(output_path)
        print(f"ICR report saved to: {output_path}")
    except Exception as e:
        print(f"Error saving ICR report: {e}")
        raise e

    return output_path
